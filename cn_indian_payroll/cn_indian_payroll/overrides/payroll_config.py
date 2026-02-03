import frappe
from openpyxl import Workbook, load_workbook
from frappe.utils.file_manager import save_file, get_file




@frappe.whitelist()
def process_attendance_regularization(payroll_entry):
    payroll_doc = frappe.get_doc("Payroll Entry", payroll_entry)

    for row in payroll_doc.custom_attendance_regularize_child:

        # ✅ Check if LOP Reversal already exists
        existing_lop = frappe.get_list(
            "LOP Reversal",
            filters={
                "employee": row.employee,
                "salary_slip": row.salary_slip,
                "payroll_period": row.payroll_period,
                "docstatus": ["!=", 2],  # not cancelled
            },
            pluck="name",
        )

        # ⏭ Skip if already created
        if existing_lop:
            continue

        # ✅ Create new LOP Reversal
        lop_doc = frappe.new_doc("LOP Reversal")
        lop_doc.employee = row.employee
        lop_doc.salary_slip = row.salary_slip
        lop_doc.attendance_log = row.attendance_log
        lop_doc.lop_month_reversal = row.month
        lop_doc.payroll_period = row.payroll_period
        lop_doc.company = payroll_doc.company
        lop_doc.additional_salary_date = row.additional_salary_date
        lop_doc.working_days = row.working_days
        lop_doc.number_of_days = row.arrear_days

        lop_doc.insert(ignore_permissions=True)
        lop_doc.submit()

    frappe.msgprint("Attendance Regularization processed successfully.")




@frappe.whitelist()
def download_new_joinee_excel(docname):
    import os
    from openpyxl import Workbook
    from frappe.utils import now_datetime

    doc = frappe.get_doc("Payroll Entry", docname)

    if not doc.custom_new_joinee_arrear_child:
        frappe.throw("No New Joinee data found")

    wb = Workbook()
    ws = wb.active
    ws.title = "New Joinee"

    ws.append([
        "Employee",
        "Employee Name",
        "Working Days",
        "Arrear Days",
        "Date Of Joining",
        "Total Earning Arrear",
        "Total Deduction Arrear",
    ])

    for row in doc.custom_new_joinee_arrear_child:
        ws.append([
            row.employee,
            row.employee_name,
            row.working_days,
            row.arrear_days,
            row.date_of_joining,
            row.total_earning_arrear,
            row.total_deduction_arrear,

        ])

    file_name = f"New Joinee_{docname}_{now_datetime().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = frappe.get_site_path("public", "files", file_name)

    wb.save(file_path)

    return f"/files/{file_name}"







@frappe.whitelist()
def download_attendance_excel(docname):
    import os
    from openpyxl import Workbook
    from frappe.utils import now_datetime

    doc = frappe.get_doc("Payroll Entry", docname)

    if not doc.custom_employee_attendance_details_list:
        frappe.throw("No attendance data found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.append([
        "Employee",
        "Employee Name",
        "Working Days",
        "Absent Days",
        "LWP Days",
        "Present Days",
        "Half Days",
        "On Leave",
        "Total LWP",
        "Payment Days"
    ])

    for row in doc.custom_employee_attendance_details_list:
        ws.append([
            row.employee,
            row.employee_name,
            row.working_days,
            row.absent_days,
            row.lwp_days,
            row.present_days,
            row.half_days,
            row.on_leave,
            row.total_lwf_days,
            row.payment_days,

        ])

    file_name = f"Attendance_{docname}_{now_datetime().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = frappe.get_site_path("public", "files", file_name)

    wb.save(file_path)

    return f"/files/{file_name}"



@frappe.whitelist()
def download_attendance_regularise_excel(docname):
    import os
    from openpyxl import Workbook
    from frappe.utils import now_datetime

    doc = frappe.get_doc("Payroll Entry", docname)

    if not doc.custom_attendance_regularize_child:
        frappe.throw("No attendance data found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    ws.append([
        "Employee",
        "Employee Name",
        "Month",
        "Working Days",
        "Arrear Days"
    ])

    for row in doc.custom_attendance_regularize_child:
        ws.append([
            row.employee,
            row.employee_name,
            row.month,
            row.working_days,
            row.arrear_days,


        ])

    file_name = f"Attendance_Regularise_{docname}_{now_datetime().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = frappe.get_site_path("public", "files", file_name)

    wb.save(file_path)

    return f"/files/{file_name}"





import frappe
import openpyxl
from frappe.utils.file_manager import get_file
from frappe.utils import getdate
import os

@frappe.whitelist()
def process_extra_payment_excel(file_url, parent_docname=None):
    """
    Read uploaded Excel and return rows for child table insertion
    """

    if not file_url:
        frappe.throw("No file attached")

    # Get File document using file_url
    file_doc = frappe.get_doc("File", {"file_url": file_url})

    # Full path on server
    full_path = frappe.get_site_path(file_doc.file_url.lstrip("/"))

    if not os.path.exists(full_path):
        frappe.throw(f"File not found at {full_path}")

    # Open Excel
    wb = openpyxl.load_workbook(full_path, data_only=True)
    sheet = wb.active

    rows = []
    header = True

    for row in sheet.iter_rows(values_only=True):
        if header:
            header = False
            continue

        if not row or not row[0]:
            continue

        rows.append({
            "employee": row[0],
            "employee_name": row[1],
            "salary_component": row[2],
            "amount": float(row[3] or 0),
            "payout_date": getdate(row[4]) if row[4] else None,
            "is_recurring": 1 if str(row[5]).lower() in ("yes", "1", "true") else 0,
            "from_date": getdate(row[6]) if row[6] else None,
            "to_date": getdate(row[7]) if row[7] else None,
            "clockback_date": getdate(row[8]) if row[8] else None,
        })

    return rows





import frappe
from openpyxl import Workbook
from io import BytesIO

@frappe.whitelist()
def download_extrapayment_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Extra Payments"

    headers = [
        "employee",
        "employee_name",
        "salary_component",
        "amount",
        "payout_date",
        "is_recurring",
        "from_date",
        "to_date",
        "clockback_date",
    ]

    ws.append(headers)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.local.response.filename = "Payroll_Extrapayment_Template.xlsx"
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "binary"



@frappe.whitelist()
def download_offcycle_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Extra Payments"

    headers = [
        "employee",
        "employee_name",
        "salary_component",
        "amount",
        "payout_date",
        "is_recurring",
        "from_date",
        "to_date",
        "clockback_date",
        "is_tax_auto_calculate",
        "is_tax_manual_calculate",
        "tds_value",
    ]

    ws.append(headers)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.local.response.filename = "Payroll_offcycle_Template.xlsx"
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "binary"



import frappe
from openpyxl import load_workbook
from io import BytesIO

@frappe.whitelist()
def process_uploaded_excel(payroll_entry):
    payroll_doc = frappe.get_doc("Payroll Entry", payroll_entry)

    if not payroll_doc.custom_upload_file:
        frappe.throw("Please upload an Excel file")

    file_doc = frappe.get_doc("File", {"file_url": payroll_doc.custom_upload_file})

    content = file_doc.get_content()
    if isinstance(content, str):
        content = content.encode()

    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    required_headers = {
        "employee",
        "employee_name",
        "salary_component",
        "amount"
    }

    if not required_headers.issubset(set(headers)):
        frappe.throw(
            "Invalid template. Required columns: "
            + ", ".join(required_headers)
        )

    # Clear existing rows (same as CSV version)
    payroll_doc.set("custom_additional_extra_payments_details", [])

    success = 0
    errors = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        row_data = dict(zip(headers, row))

        try:
            payroll_doc.append(
                "custom_additional_extra_payments_details",
                {
                    "employee": row_data.get("employee"),
                    "employee_name": row_data.get("employee_name"),
                    "salary_component": row_data.get("salary_component"),
                    "amount": row_data.get("amount"),
                    "payout_date": row_data.get("payout_date"),

                    "is_recurring": int(row_data.get("is_recurring") or 0),
                    "from_date": row_data.get("from_date"),
                    "to_date": row_data.get("to_date"),
                    "clockback_date": row_data.get("clockback_date"),
                }
            )
            success += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    payroll_doc.save(ignore_permissions=True)

    return {
        "success": success,
        "errors": errors
    }


@frappe.whitelist()
def process_uploaded_excel_offcycle(payroll_entry):
    payroll_doc = frappe.get_doc("Payroll Entry", payroll_entry)

    if not payroll_doc.custom_offcycle_attach:
        frappe.throw("Please upload an Excel file")

    file_doc = frappe.get_doc("File", {"file_url": payroll_doc.custom_offcycle_attach})

    content = file_doc.get_content()
    if isinstance(content, str):
        content = content.encode()

    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    required_headers = {
        "employee",
        "employee_name",
        "salary_component",
        "amount"
    }

    if not required_headers.issubset(set(headers)):
        frappe.throw(
            "Invalid template. Required columns: "
            + ", ".join(required_headers)
        )

    # Clear existing rows (same as CSV version)
    payroll_doc.set("custom_offcycle_data", [])

    success = 0
    errors = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        row_data = dict(zip(headers, row))

        try:
            payroll_doc.append(
                "custom_offcycle_data",
                {
                    "employee": row_data.get("employee"),
                    "employee_name": row_data.get("employee_name"),
                    "salary_component": row_data.get("salary_component"),
                    "amount": row_data.get("amount"),
                    "payout_date": row_data.get("payout_date"),

                    "is_recurring": int(row_data.get("is_recurring") or 0),
                    "from_date": row_data.get("from_date"),
                    "to_date": row_data.get("to_date"),
                    "clockback_date": row_data.get("clockback_date"),
                    "is_tax_auto_calculate":row_data.get("is_tax_auto_calculate"),
                    "is_tax_manual_calculate":row_data.get("is_tax_manual_calculate"),
                    "tds_value":row_data.get("tds_value")
                }
            )
            success += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    payroll_doc.save(ignore_permissions=True)

    return {
        "success": success,
        "errors": errors
    }







import frappe
from frappe.utils import today

@frappe.whitelist()
def create_extra_payment(docname):
    payroll_entry = frappe.get_doc("Payroll Entry", docname)

    if not payroll_entry.custom_additional_extra_payments_details:
        frappe.throw("No extra payment rows found")

    created = 0
    skipped = 0

    for row in payroll_entry.custom_additional_extra_payments_details:

        # 1️⃣ Skip invalid rows
        if not row.employee or not row.salary_component or not row.amount:
            continue

        payroll_date = row.payout_date or row.from_date

        # 2️⃣ DB-level exists check (employee + component + payroll_date)
        exists = frappe.db.exists(
            "Additional Salary",
            {
                "employee": row.employee,
                "salary_component": row.salary_component,
                "payroll_date": payroll_date,
                "docstatus": ["!=", 2],  # ignore cancelled
            }
        )

        if exists:
            skipped += 1

            # mark as processed (optional but recommended)
            if not row.payout_date:
                row.payout_date = payroll_date

            continue

        # 3️⃣ Create Additional Salary
        additional_salary = frappe.new_doc("Additional Salary")
        additional_salary.employee = row.employee
        additional_salary.salary_component = row.salary_component
        additional_salary.amount = row.amount

        additional_salary.ref_doctype = "Payroll Entry"
        additional_salary.ref_docname = payroll_entry.name

        additional_salary.payroll_date = payroll_date
        additional_salary.is_recurring = row.is_recurring or 0
        additional_salary.from_date = row.from_date
        additional_salary.to_date = row.to_date

        additional_salary.insert(ignore_permissions=True)
        additional_salary.submit()

        # 4️⃣ Mark row as processed
        row.payout_date = payroll_date

        created += 1

    payroll_entry.save(ignore_permissions=True)
    frappe.db.commit()

    return {
        "created": created,
        "skipped": skipped
    }




@frappe.whitelist()
def create_offcycle_payment(docname):
    payroll_entry = frappe.get_doc("Payroll Entry", docname)

    if not payroll_entry.custom_offcycle_data:
        frappe.throw("No extra payment rows found")

    created = 0
    skipped = 0

    for row in payroll_entry.custom_offcycle_data:

        if not row.employee or not row.salary_component or not row.amount:
            continue

        payroll_date = row.payout_date or row.from_date


        exists = frappe.db.exists(
            "Additional Salary",
            {
                "employee": row.employee,
                "salary_component": row.salary_component,
                "payroll_date": payroll_date,
                "docstatus": ["!=", 2],  # ignore cancelled
            }
        )

        if exists:
            skipped += 1

            # mark as processed (optional but recommended)
            if not row.payout_date:
                row.payout_date = payroll_date

            continue

        additional_salary = frappe.new_doc("Additional Salary")
        additional_salary.employee = row.employee
        additional_salary.salary_component = row.salary_component
        additional_salary.amount = row.amount

        additional_salary.ref_doctype = "Payroll Entry"
        additional_salary.ref_docname = payroll_entry.name

        additional_salary.payroll_date = payroll_date
        additional_salary.is_recurring = row.is_recurring or 0
        additional_salary.from_date = row.from_date
        additional_salary.to_date = row.to_date

        additional_salary.deduct_full_tax_on_selected_payroll_date = row.is_tax_auto_calculate
        additional_salary.custom_is_tax_manual_calculate = row.is_tax_manual_calculate
        additional_salary.custom_tds_value = row.tds_value


        additional_salary.insert(ignore_permissions=True)
        additional_salary.submit()

        row.payout_date = payroll_date

        created += 1

    payroll_entry.save(ignore_permissions=True)
    frappe.db.commit()

    return {
        "created": created,
        "skipped": skipped
    }
